import openpyxl
import sqlite3 as sq


def take_from_database():
    numbers_from_sql = []
    messages_from_sql = []
    with sq.connect('dates.db') as con:
        cur = con.cursor()
        cur.execute("""SELECT numbers , message FROM bot_voisa_msk""")
        for a in cur:
            number, message = a
            numbers_from_sql.append(number)
            messages_from_sql.append(message)
    return numbers_from_sql, messages_from_sql


def take_from_database_2():
    numbers_from_sql = []
    messages_from_sql = []
    with sq.connect('dates.db') as con:
        cur = con.cursor()
        cur.execute("""SELECT numbers , message FROM bot_voisa_spb""")
        for a in cur:
            number, message = a
            numbers_from_sql.append(number)
            messages_from_sql.append(message)
    return numbers_from_sql, messages_from_sql


def xlsx_from_msk():   # Создает и заполняет excel файл
    data = take_from_database()
    column_A, column_B = data
    book = openpyxl.Workbook()
    sheet = book.create_sheet('All_numbers_msk')
    book.remove(book.active)
    for i, name in enumerate(column_A):
        sheet[f'A{i+1}'] = name
    for i, name in enumerate(column_B):
        sheet[f'B{i+1}'] = name
    book.save('numbers_msk.xlsx')
    book.close()


def xlsx_from_spb():   # Создает и заполняет excel файл
    data = take_from_database_2()
    column_A, column_B = data
    book = openpyxl.Workbook()
    sheet = book.create_sheet('All_numbers_spb')
    book.remove(book.active)
    for i, name in enumerate(column_A):
        sheet[f'A{i+1}'] = name
    for i, name in enumerate(column_B):
        sheet[f'B{i+1}'] = name
    book.save('numbers_spb.xlsx')
    book.close()
